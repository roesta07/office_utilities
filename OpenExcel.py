from openpyxl import load_workbook,worksheet
import pandas as pd
import datetime as dt
import os
import re
from itertools import chain
from openpyxl.utils.cell import get_column_letter

class OpenExcel:
    """
    input: Excel filepath with table name
    return: pandas dataframe object
            -concatenate tables if there are multiple files in the folder

    For Example:
    -   df=OpenExcel(filename='party_ledger_v1.1.xlsx').from_table('ledger_data')
    -   df=OpenExcel().from_folder(ROOT_DIR,'VatBills_Sales')
    """
    def __init__(self,filename=None,no_formula=True):  
        if filename is not None:
            self.filename=filename
            self.wb=load_workbook(filename=self.filename,data_only=no_formula) ## data_only argument so that we dont get excel formulas
            self.mapping={self.wb:{}}
            for ws in self.wb.worksheets:
                for tbl_name,ref in ws.tables.items():
                    self.mapping[self.wb][tbl_name]=[ws,ref]

    def filter_files(self,name):
        filter_excel=(name.endswith('.xlsx')) & (not name.startswith('~')) ##filter file name and temp
        filter_conflict=not bool(re.search('conflict',name))
        return filter_excel & ( filter_conflict)
    
    def get_table_data(self,workbook,table_name):
        try:
            if self.mapping[workbook][table_name]:
                table_sheet=self.mapping[workbook][table_name][0]
                table_ref=self.mapping[workbook][table_name][1]
                data=table_sheet[table_ref] ## returns all the cell reference where each row is a tuple
                return data
        except:
            print('No table with that Name found')
      
    def from_table(self,table_name): 
        data=self.get_table_data(self.wb,table_name)
        content=[[cell.value for cell in ent] for ent in data] ## value for each table cells; list compression method
        header=content[0] ## first row as header
        rest=content[1:] ## rest rows as content
        df=pd.DataFrame(rest,columns=header) ## converting into dataframe
        return df

    def update_mappings(self,root_path):
        files={root_dir_path:list(filter(self.filter_files,filename)) 
               for root_dir_path,sub_dir, filename in os.walk(root_path) 
               if filename}        
        nested_filelist=[[os.path.join(k,v) for v in vs if v] for k,vs in files.items()]
        filelist = list(chain.from_iterable(nested_filelist))
        self.wbs=[load_workbook(filename=item,data_only=True) for item in filelist]
        self.mapping={wb:{} for wb in self.wbs}
        for wb in self.wbs:
            for ws in wb.worksheets:
                for tbl_name,ref in ws.tables.items():
                    self.mapping[wb][tbl_name]=[ws,ref]
          
    def from_folder(self,root_path,tablename):
        self.update_mappings(root_path)
        dfs=[]
        for wb in self.mapping.keys():
            data=self.get_table_data(wb,tablename)
            content=[[cell.value for cell in ent] for ent in data] ## value for each table cells; list compression method
            header=content[0] ## first row as header
            rest=content[1:] ## rest rows as content
            df=pd.DataFrame(rest,columns=header) ## converting into dataframe
            dfs.append(df)
            dataframe=pd.concat(dfs)
        return dataframe
        
# df=OpenExcel(filename='party_ledger_v1.1.xlsx').from_table('ledger_data')
# files=OpenExcel().from_folder(ROOT_DIR,'VatBills_Sales')