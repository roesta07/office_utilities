import shutil
import os
import sys
from datetime import datetime as dt
from openpyxl import load_workbook
import numpy as np
import json
import pandas as pd


class Migrate:
    CURRENT=os.getcwd()

    BASE_TARGET='C:\\Dropbox\\Dropbox\\PTK Nepal\\Operation'
    cash_matches=('Bot','Binay','anjila')
    def __init__(self):
        
        self.NAMES=[] ## names in folder
        self.ORIGINAL=[]
        self.TARGET=[]
        ## Bot functions
        self.update()
    
    ## utility functions 
    def filter_files(self,x):
        """
        utility function that runs itself while running process
        function output: filter excel files in the directory
        """
        if x.endswith('xlsx'):
            return True
        else:
            return False
    #utility function   
    def assign_root(self,x):
        """
        Utility function:
        Assign root: Where each file has to go inside Entry_users
        Cases Handled>>
        -Attendance 
        -DCbooks
        -cashflows* (depends on cash_matches)  
        """
        if 'Attendance' in x:
            ROOT=f'Recruitment\\#EntryAttendance\\{x}'
            return os.path.join(Migrate.BASE_TARGET,ROOT)
        elif 'DCBooks' in x:
            fy=x.split('_')[1]
            ROOT=f'Finance\\#Entry_users\\{fy}\\{x}'
            return os.path.join(Migrate.BASE_TARGET,ROOT)
    
        elif any(match in x for match in Migrate.cash_matches):
            """
            for cash_matches
            """
            fy=x.split('_')[1]
            return os.path.join(Migrate.BASE_TARGET,f'Finance\\#Entry_users\\{fy}\\{x}')
        
    def update(self):
        """
        This updates where to put the files from where(default=current directory)
        """
        FILENAMES=[filename for root_dir_path,sub_dir, filename in os.walk(Migrate.CURRENT)][0]
        self.NAMES=list(filter(self.filter_files,FILENAMES))
        print(self.NAMES)
        self.ORIGINAL=[os.path.join(Migrate.CURRENT,FILE) for FILE in self.NAMES]
        self.TARGET=list(map(self.assign_root,self.NAMES))
        

    def getParams(self):
        """Returns self.ORIGINAL and Self.TARGET """
        return self.ORIGINAL,self.TARGET

## other utilities
##utility function
def filter_files(x):
        if x.endswith('xlsx') and '~' not in x:
            return True
        if '~' in x or 'conflict' in x:
            print('Program Ran But Some files are still Open or conflicted; \nSave and close the program to get full affect\nignore this message incase of migration')
        else:
            return False

def main():
    def load_rename_file(cwd=None,filename='find_replace.csv'):
        cwd=os.getcwd()
        filepath=os.path.join(cwd,filename)
        df=pd.read_csv(filepath)
        renames={k:v for k,v in zip(df['find'],df['replace'])}

        return renames

    def help():
        print(f'run bot.py with: \n--migrate\n--backup\n--backupandmigrate')

    def migrate():
        """
        This function is called without any params;
        When called: Migrates every files from backup folders to target folders
        classify

        """
        print('migrating')
        MM=Migrate()
        original_paths,target_paths=MM.getParams()
        
        for original,target in zip(original_paths,target_paths):
         
            shutil.copyfile(original,target)


    def backup():
        EXTRACT_FROM_BASE=['C:\\Dropbox\\Dropbox\\PTK Nepal\\Operation\\Recruitment\\#EntryAttendance',
            'C:\\Dropbox\\Dropbox\\PTK Nepal\\Operation\\Finance\\#Entry_users',
            'C:\\Dropbox\\Dropbox\\PTK Nepal\\Operation\\Finance\\DebtsCredits_Book']
        BASE_BACKUP='C:\\Dropbox\\Dropbox\\PTK Nepal\\Operation\\Backups' 
        
        ## FOR BACKUP
        extracts=[]
        extract_to_root=[]
        now=dt.now()
        folder_name = now.strftime("%Y-%b-%d_%H-%M-%S")
        os.mkdir(os.path.join(BASE_BACKUP,folder_name))
        for path in EXTRACT_FROM_BASE:
            for root_dir_path,sub_dir, filenames  in os.walk(path):
                if filenames:
                        for filename in filenames:
                            extracts.append(os.path.join(root_dir_path,filename))
                            extract_to_root.append(os.path.join(BASE_BACKUP,folder_name,filename))
        EXTRACT_NAMES=list(filter(filter_files,extracts))
        EXTRACT_TO=list(filter(filter_files,extract_to_root))
        for original,target in zip(EXTRACT_NAMES,EXTRACT_TO):
                shutil.copyfile(original,target)


    def rename(renames=None):
        """
        input: renames(filepath).csv
        output: find and replace every files in the directory
        """
        renames=load_rename_file()
        if renames==None:
            print('Please provide Rename values')

        else:
            ## list all the excel_files from the directory
            wb_filenames=list(filter(filter_files,os.listdir()))

            
            for wb_filename in wb_filenames: ## loop workbooks
                wb=load_workbook(filename=wb_filename,data_only=True)
                mapping={wb:{}}
                for ws in wb.worksheets:
                    for tbl_name,ref in ws.tables.items():
                        mapping[wb][tbl_name]=[ws,ref]

                
                for k in mapping[wb].keys():
                    ##get worksheet
                    ws=mapping[wb][k][0]
                    table_ref=mapping[wb][k][1]
                    data=ws[table_ref]
                    ## to make the loop easier lets stack them all in one place
                    cell_ref=np.hstack([[f'{cell.column_letter}{cell.row}' for cell in ent]for ent in data])
                    cell_value=np.hstack([[cell.value for cell in ent]for ent in data])

                    for key in renames.keys():
                        for ref,value in zip(cell_ref,cell_value):
                            if value==key:
                                ws[ref]=renames[key]

                wb.save(wb_filename)
    if sys.argv[1]=='help':
        help()
    if sys.argv[1]=='migrate':
        print("****Initiating..........Please wait")
        print('*****running migrations*****')
        migrate()
        print('*****Migrated successfully*****')
    if sys.argv[1]=='backup':
        print('*****Backing up files*****')
        backup()
    if sys.argv[1]=='backupandmigrate':
        print("****Initiating..........Please wait")
        backup()
        migrate()
        print('*****Successfull:Backup baand Migrations*****')
    if sys.argv[1]=='rename':
        renames=load_rename_file()
        rename(renames)
    else:
        help()

if __name__ == '__main__':
    main()